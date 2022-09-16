import tkinter as tk

root = tk.Tk()

v = tk.IntVar()


from openpyxl import *
from tkinter import *
wb = load_workbook('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
sheet = wb.active
def excel():
    sheet.column_dimensions['G'].width = 30
    sheet.cell(row=1, column=7).value = "Ambulance"
excel()


def oneone():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=7).value = "Yes"
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
def twotwo():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=7).value = "No"
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')

def lastpage():
    import lastpage.py
    exec(lastpage.py)

tk.Label(root,
        text="""Ambulance needed:""",
        justify = tk.LEFT,
        padx = 20).pack()
tk.Radiobutton(root,
              text="YES",
              padx = 20,
              variable=v,
              value=1, command= oneone).pack(anchor=tk.W)
tk.Radiobutton(root,
              text="NO",
              padx = 20,
              variable=v,
              value=2, command= twotwo).pack(anchor=tk.W)

hey= tk.Button(root, text="Next", fg="Black",
                            bg="Red", command= lastpage)
hey.pack()

    



root.mainloop()


    



