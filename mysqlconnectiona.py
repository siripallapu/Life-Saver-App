import mysql.connector
mydb= mysql.connector.connect(host= 'localhost', database= 'server2', user= 'root', password='siri')

mycursor = mydb.cursor()
from openpyxl import *
from tkinter import *
wb = load_workbook('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
sheet = wb.active


oneone=  sheet.cell(row=sheet.max_row, column=1).value
twotwo=  sheet.cell(row=sheet.max_row, column=2).value
three=  sheet.cell(row=sheet.max_row, column=3).value
four=  sheet.cell(row=sheet.max_row, column=4).value
five=  sheet.cell(row=sheet.max_row, column=5).value
six=  sheet.cell(row=sheet.max_row, column=6).value
seven=  sheet.cell(row=sheet.max_row, column=7).value

hos= sheet.cell(row= sheet.max_row, column=8).value
case=  sheet.cell(row=sheet.max_row , column=5).value

if hos== "hospital 2" and case=="Emergency":
    sql = "INSERT INTO emergency (Name, contactno, emailid, address, casetype, casename, ambulanceneeded) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = (oneone, twotwo, three, four, five, six, seven)
    mycursor.execute(sql, val)
elif hos== "hospital 2" and case=="Nonemergency":
    sql = "INSERT INTO nonemergency (Name, contactno, emailid, address, casetype, casename, ambulanceneeded) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = (oneone, twotwo, three, four, five, six, seven)
    mycursor.execute(sql, val)
elif hos== "hospital 2" and case=="Covid19":
    sql = "INSERT INTO covid19 (Name, contactno, emailid, address, casetype, casename, ambulanceneeded) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = (oneone, twotwo, three, four, five, six, seven)
    mycursor.execute(sql, val)
elif hos== "hospital 2" and case=="other" :
    sql = "INSERT INTO other(Name, contactno, emailid, address, casetype, casename, ambulanceneeded) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = (oneone, twotwo, three, four, five, six, seven)
    mycursor.execute(sql, val)
    
    
    
    


mydb.commit()

print(mycursor.rowcount, "record inserted.")


