

import tkinter as tk
from tkinter import ttk
     
root = tk.Tk()

v = tk.IntVar()
v.set(0)  # initializing the choice, i.e. Python


        
tk.Label(root,
         text="""Choose the Problem""",
         justify = tk.LEFT,
         padx = 20).pack()

from openpyxl import *
from tkinter import *
wb = load_workbook('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
sheet = wb.active
def excel():
    sheet.column_dimensions['H'].width = 30
    sheet.cell(row=1, column=6).value = "Problem"
excel()
     
current_row = sheet.max_row
current_column = sheet.max_column
sheet.cell(row=current_row + 1, column=5).value = "Covid19"

def oneone():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=6).value = "Complication"
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
def twotwo():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=6).value = "Checkup"
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
def threethree():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=6).value = "Other"
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')

        

tk.Radiobutton(root,
                text="Complication",
                indicatoron = 0,
                width = 20,
                padx = 20,
                variable=v,
                command= oneone,
                ).pack(anchor=tk.W)

tk.Radiobutton(root,
                text="Checkup",
                indicatoron = 0,
                width = 20,
                padx = 20,
                variable=v,
                command= twotwo,
                ).pack(anchor=tk.W)

tk.Radiobutton(root,
                text="Other",
                indicatoron = 0,
                width = 20,
                padx = 20,
                variable=v,
                command= threethree,
                ).pack(anchor=tk.W)
    

def location():
    import webbrowser
    webbrowser.open("maps.html")
        
show = tk.Button(root, text="see locations of hospitals", fg="Black",
                            bg="Red", command= location)

show.pack()

def ch_hos():
    import chosehospital.py
    exec(chosehospital.py)

choose= tk.Button(root, text="choose hospital", fg="Black",
                            bg="Red", command= ch_hos)
choose.pack()
     
 
root.mainloop()

    


