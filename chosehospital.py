import tkinter as tk

root = tk.Tk()

v = tk.IntVar()
v.set(0)  # initializing the choice, i.e. Python


        
tk.Label(root,
         text="""Choose the Hospital""",
         justify = tk.LEFT,
         padx = 20).pack()

from openpyxl import *
from tkinter import *
wb = load_workbook('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
sheet = wb.active
def excel():
    sheet.column_dimensions['H'].width = 30
    sheet.cell(row=1, column=8).value = "Hospital Chosen"
excel()


def savehospital1():
    hello= "hospital 1"
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=8).value = hello
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
    

def savehospital2():
    hello= "hospital 2"
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=8).value = hello
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')

def savehospital3():
    hello= "hospital 3"
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row, column=8).value = hello
    wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx') 
    
    
    

tk.Radiobutton(root,
                text="Hospital 1",
                indicatoron = 0,
                width = 20,
                padx = 20,
                variable=v,
                command= savehospital1,
                ).pack(anchor=tk.W)

tk.Radiobutton(root,
                text="Hospital 2",
                indicatoron = 0,
                width = 20,
                padx = 20,
                variable=v,
                command= savehospital2,
                ).pack(anchor=tk.W)

tk.Radiobutton(root,
                text="Hospital 3",
                indicatoron = 0,
                width = 20,
                padx = 20,
                variable=v,
                command= savehospital3,
                ).pack(anchor=tk.W)
def formm():
    import form.py
    exec("C:\\Users\\USER\\Desktop\\siri\\computer_project_completed_files\\form.py")
nexxt = Button(root, text="Next", fg="Black",
                            bg="Red", command= formm)
nexxt.pack()
root.mainloop()

