

from openpyxl import *
from tkinter import *
wb = load_workbook('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
sheet = wb.active
def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Contact Number"
    sheet.cell(row=1, column=3).value = "Email ID"
    sheet.cell(row=1, column=4).value = "Address"
def focus1(event):
    contact_no_field.focus_set()
def focus2(event):
    email_id_field.focus_set()
def focus3(event):
    address_field.focus_set()
def clear():
    name_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)
def insert():
    if (name_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):
        print("empty input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
        sheet.cell(row=current_row, column=1).value = name_field.get()
        sheet.cell(row=current_row, column=2).value = contact_no_field.get()
        sheet.cell(row=current_row, column=3).value = email_id_field.get()
        sheet.cell(row=current_row, column=4).value =address_field.get()  
        wb.save('C:\\Users\\USER\\Desktop\\lifesaver database.xlsx')
        name_field.focus_set()
        clear()
def ambulance():
    import ambulance.py
    exec(ambulance.py)

    
    
if __name__ == "__main__":
    root = Tk()
    root.configure(background='light green')
    root.title("lifesaver form")
    root.geometry("500x300")
 
    excel()
    heading = Label(root, text="Form", bg="light green")
    name = Label(root, text="Name", bg="light green")
    contact_no = Label(root, text="Contact No.", bg="light green")  
    email_id = Label(root, text="Email id", bg="light green")
    address = Label(root, text="Address", bg="light green")
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    contact_no.grid(row=2, column=0)
    email_id.grid(row=3, column=0)
    address.grid(row=4, column=0)
    name_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)
    name_field.bind("<Return>", focus1)
    contact_no_field.bind("<Return>", focus2)
    email_id_field.bind("<Return>", focus3)
    name_field.grid(row=1, column=1, ipadx="100")
    contact_no_field.grid(row=2, column=1, ipadx="100")
    email_id_field.grid(row=3, column=1, ipadx="100")
    address_field.grid(row=4, column=1, ipadx="100")
    excel()
    submit = Button(root, text="Submit", fg="Black",
                            bg="Red", command= insert)
    submit.grid(row=5, column=1)

    nexxt = Button(root, text="Next", fg="Black",
                            bg="Red", command= ambulance)
    nexxt.grid(row=6, column=1)
    root.mainloop()


